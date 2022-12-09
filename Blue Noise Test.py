import openpyxl
import random
import matplotlib.pyplot as plt
import math

visitedRows = set()

screen_width = 640
screen_height = 480


def torDist(firstPt, secondPt):
    x1 = firstPt[0]
    y1 = firstPt[1]

    x2 = secondPt[0]
    y2 = secondPt[1]
    
    xdiff = abs(x2 - x1)
    if xdiff > (screen_width / 2):
        xdiff = screen_width - xdiff

    ydiff = abs(y2 - y1)
    if ydiff > (screen_height / 2):
        ydiff = screen_height - ydiff

    return math.sqrt(xdiff**2 + ydiff**2)

def blueNoiseSample():
    wrkbk = openpyxl.load_workbook("heightWeightChart.xlsx")
    sh = wrkbk.active

    minX, maxX = 1000000000, -1000000000
    minY, maxY = 1000000000, -1000000000

    desiredSamplePts = 250

    '''
    for i in range(2, sh.max_row + 1):
        xVal = sh.cell(row = i, column = 2).value
        
        if float(xVal) < minX:
            minX = float(xVal)

        if float(xVal) > maxX:
            maxX = float(xVal)

        yVal = sh.cell(row = i, column = 3).value
        
        if float(yVal) < minY:
            minY = float(yVal)

        if float(yVal) > maxY:
            maxY = float(yVal)
    '''
    
    m = 5
    currSamplePts = 1

    pointsToGraph = []
    
    randfirstRow = random.randint(2, sh.max_row)

    xValRandFirstPt = sh.cell(row = randfirstRow, column = 2).value
    yValRandFirstPt = sh.cell(row = randfirstRow, column = 3).value
    
    pointsToGraph.append((float(xValRandFirstPt), float(yValRandFirstPt)))

    while currSamplePts < desiredSamplePts:
        print(currSamplePts)

        lastSamplePt = pointsToGraph[-1]
        numPtsToGen = currSamplePts * m

        tempPtsList = []
        for i in range(numPtsToGen):

            randRow = random.randint(2, sh.max_row)
            xVal = sh.cell(row = randRow, column = 2).value
            yVal = sh.cell(row = randRow, column = 3).value
            
            
            tempPtsList.append((float(xVal), float(yVal)))

        #Calculate minimalToroidDist between every point in tempPtsList and lastSamplePt
        #pointsToGraph.append(maxTorDistXCoord, maxTorDistYCoord)

        candPt = (0, 0)
        maxDist = -5

        for point in tempPtsList:
            if torDist(point, lastSamplePt) > maxDist:
                maxDidst = torDist(point, lastSamplePt)
                candPt = point

        pointsToGraph.append(candPt)
        currSamplePts += 1
        

        
    height, weight = zip(*pointsToGraph)
    plt.scatter(height, weight)
    plt.show()
        
    
    print("Donezo: ")
        
        
        
blueNoiseSample()
